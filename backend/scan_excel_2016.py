# -*- coding: utf-8 -*-
import sys
sys.path.insert(0, '.')
import openpyxl

wb = openpyxl.load_workbook(r'data/uploads/cbs_2016.xlsx', read_only=True, data_only=True)

with open('excel_2016_scan.txt', 'w', encoding='utf-8') as f:
    f.write(f'Sheets: {wb.sheetnames}\n\n')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f'=== Sheet: {sheet_name} ===\n')
        rows = list(ws.iter_rows(values_only=True))
        f.write(f'Rows: {len(rows)}, Cols: {ws.max_column}\n')
        # print first 10 rows
        for i, row in enumerate(rows[:10]):
            f.write(f'  row {i}: {row[:8]}\n')
        f.write('\n')

        # search for נחל שורק / עמק לוד in all rows
        targets = ['נחל שורק', 'עמק לוד', 'נחל', 'עמק לוד']
        found_rows = []
        for i, row in enumerate(rows):
            for cell in row:
                if cell and isinstance(cell, str):
                    if any(t in cell for t in ['נחל שורק', 'עמק לוד']):
                        found_rows.append((i, row))
                        break
        if found_rows:
            f.write(f'  FOUND TARGET ROWS:\n')
            for i, row in found_rows:
                f.write(f'    row {i}: {row}\n')
        else:
            f.write(f'  (not found in this sheet)\n')
        f.write('\n')

wb.close()
print('Done - see excel_2016_scan.txt')
